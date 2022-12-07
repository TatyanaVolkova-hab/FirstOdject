#ver2
from pprint import pprint
from bs4 import BeautifulSoup               #для распознавания html
from rutermextract import TermExtractor     #для распознавания ключевых фраз
import openpyxl                             #для работы с экселем         
from tkinter import *                       #для интерфейса                     
import requests as req                      #для работы с запросами
import urllib.request
import os

WORDS=['тип антенны', 'диапазон частот', 'выходная мощность', 'мощность на входе', 
       'высота установки', 'азимут', 'тип модуляции', 'усиления', 'ширина ДН']

#функция для открытия файла экселя (после распознавания)
def open_table():
    os.startfile('table.xlsx')

def openSZZ():
    s=[]
    #получаем текст заключения от сервера
    resp = req.get(txt.get())
    #распознаем в нем теги
    soup = BeautifulSoup(resp.text, 'lxml')
    #открываем эксель для записи
    wb = openpyxl.load_workbook('table.xlsx')
    sheet = wb.active
    term_extractor = TermExtractor()
    #назначаем список для хранения характеристик
    PRTO = [[]]
    #внутри тега <p> ищем нужные фразы (характеристики)
    for tag in soup.find_all("p"):
        n=0
        #достаем название объекта
        if tag.text.find("№")>0:
                bs_name = tag.text
                print(bs_name)
        #достаем характеристики
        if tag.text.find("передающ")>0:
            t = tag.text.lower()
            for i in t.split('.-'):
                PRTO.append([])
                PRTO[n].append({'ant name': ant_name(i)})
                PRTO[n].append({'freq': freq(i)})
                PRTO[n].append({'mod': modulation(i)})
                #PRTO[n].append({'out_power': out_power(i)})
                #for j in i.split('. '):
                for j in i.split('. '):
                    for k in j.split(', '):
                        #print(k)
                        for term in term_extractor(k):
                            if term.normalized.find('выход')>-1 or term.normalized.find('мощность передатчика')>-1 or term.normalized.find('Мощность каждого передатчика по секторам')>-1:
                            #    s.append(k[k.find('Вт')-4:k.find('Вт')-1])
                                PRTO[n].append({'out_power': k[k.find('Вт')-4:k.find('Вт')-1]})
                            if term.normalized.find('вход')>-1:
                                #записываем значение в ячейку экселя
                                sheet.cell(row = n+2, column = 5).value = j[j.find('мощность на входе'):]
                                #записываем значение в характеристики ПРТО
                                PRTO[n].append({'in_power': k[k.find('Вт')-4:k.find('Вт')-1]})
                            if term.normalized.find('высота')>-1:
                                sheet.cell(row = n+2, column = 6).value = j[j.find('высота подвеса'):]
                                #PRTO[n]['H'] = i[i.find('высота подвеса'):]
                                pass
                            if term.normalized.find('азимут')>-1:
                                sheet.cell(row = n+2, column = 7).value = j[j.find('азимут'):]
                                pass
                            if term.normalized.find('ширина ДН')>-1:
                                sheet.cell(row = n+2, column = 9).value = j[j.find('ширина ДН'):]
                                pass
                        #if k.find('модуляция')>-1:
                            #sheet.cell(row = n, column = 10).value = j[j.find('модуляция'):]
                            #PRTO[n].append({'mod': k[k.find('-')+1:k.find('.')-1]})
                        if k.find('усиления')>-1:
                            sheet.cell(row = n+2, column = 8).value = j[j.find('усиления'):]
                            PRTO[n].append({'Gain': k[k.find('дБи')-8:k.find('дБи')-3]})
                n+=1
    print(s)
    pprint(PRTO)
    #сохраняем эксель с данными
    wb.save('table.xlsx')

def BS_name(text):
    poz_0 = text.find('№')
    poz_1 = text.find(' ', poz_0+2)
    return text[poz_0+1:poz_1]

ant_name_words = [
    'тип антенны', 
    'антенны типа'
]
def ant_name(text):
    for words in ant_name_words:
        if words in text:
            poz_0 = text.find(words) + len(words)
            poz_1 = text.find(' ', poz_0+2)
            poz_2 = text.find(' ', poz_1+2)
            return text[poz_0+1:poz_2]

def freq(text):
    poz_0 = text.find('диапазон частот') + 16
    poz_1 = text.find(' ', poz_0+2)
    return text[poz_0:poz_1]

out_pow_words = [
    'мощность каждого передатчика по секторам',
    'мощность передатчика'
]
def out_power(text):
    for word in out_pow_words:
        if word in text:
            return text[text.find('Вт')-4:text.find('Вт')-1]
        else:
            pass
            
def modulation(text):
    poz_0 = text.find('модуляции -') + 12
    poz_1 = text.find('.', poz_0+2)    
    if 'модуляции' in text:
        return text[poz_0:poz_1]

def to_excel(prto):
    #ws = wb.create_sheet(BS_name(bs_name))
    sheet.cell(row = n, column = 1).value = ant_name(i)
    sheet.cell(row = n, column = 2).value = freq(i)
    pass

#работа с интерфейсом
#создаем окошко
root=Tk()
root.title("Реестр СЗЗ")
root.geometry("1000x500")

    
#Кнопочки - надписи
#but1=Button(root,text="Загрузить данные реестра",command = download) 
#but1.grid(column=0, row=0)
butsearch=Button(root,text="Выполнить анализ",command = openSZZ)
butsearch.grid(column=0, row=3)
lbl1 = Label(root, text="Введите URL-адрес заключения:")  
lbl1.grid(column=0, row=0) 
txt = Entry(root,width=100)
txt.grid(column=1, row=0)
lbl = Label(root, text="Ответ сервера")  
lbl.grid(column=0, row=2) 
butopen=Button(root,text="Посмотреть результат",command = open_table)
butopen.grid(column=0, row=4)
root.mainloop()
